import pandas as pd
import numpy as np
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CLR_HEADER   = "1F4E79"
CLR_SUBHDR   = "2E75B6"
CLR_WARN     = "C00000"
CLR_CAUTION  = "BF8F00"
CLR_OK       = "375623"
CLR_NODATA   = "808080"
CLR_ROW_A    = "DCE6F1"
CLR_ROW_B    = "FFFFFF"
CLR_ASSUMPTION = "FFF2CC"

LIMIT_100 = 10
LIMIT_70  = 150
LIMIT_20  = 10000

TM_MAP = {'СКУ1': 'ТМ-1', 'СКУ3': 'ТМ-3', 'СКУ5': 'ТМ-5', 'СКУ6': 'ТМ-6', 'СКУ7': 'ТМ-7'}
TEMP_COLS = {
    'ТМ-1': ('TM1_max', 'TM1_min'),
    'ТМ-3': ('TM3_max', 'TM3_min'),
    'ТМ-4': ('TM4_max', 'TM4_min'),
    'ТМ-5': ('TM5_max', 'TM5_min'),
}


def count_seasonal_cycles(series, level):
    """Цикл = переход порога снизу вверх, затем сверху вниз (туда-обратно)."""
    s = series.dropna()
    if len(s) < 2:
        return 0
    above = (s > level).values
    cycles = 0
    crossed_up_pending = bool(above[0])
    was_above = bool(above[0])
    for cur in above[1:]:
        cur = bool(cur)
        if not was_above and cur:
            crossed_up_pending = True
        elif was_above and not cur and crossed_up_pending:
            cycles += 1
            crossed_up_pending = False
        was_above = cur
    return cycles


def find_seasonal_cycle_periods(series, level):
    """
    Возвращает список периодов (дата_начала_превышения, дата_возврата) —
    то есть когда именно температура поднялась выше порога и когда
    опустилась обратно ниже него (завершённый цикл).
    """
    s = series.dropna()
    if len(s) < 2:
        return []
    above = (s > level).values
    dates = s.index
    periods = []
    up_date = dates[0] if above[0] else None
    crossed_up_pending = bool(above[0])
    was_above = bool(above[0])
    for i in range(1, len(above)):
        cur = bool(above[i])
        if not was_above and cur:
            up_date = dates[i]
            crossed_up_pending = True
        elif was_above and not cur and crossed_up_pending:
            periods.append((up_date, dates[i]))
            crossed_up_pending = False
            up_date = None
        was_above = cur
    return periods


def count_daily_cycles_carryover(dT_series, threshold):
    """Суточный разброс с накоплением остатка (carry-over, без сброса)."""
    total, carry = 0, 0.0
    for dt in dT_series:
        available = carry + dt
        n = int(available // threshold)
        total += n
        carry = available - n * threshold
    return total


def load_temperatures(path):
    df = pd.read_excel(path, sheet_name=0, header=None)
    temp = df.iloc[2:, [0,1,2,3,4,5,6,7,8]].copy()
    temp.columns = ['date','TM1_max','TM1_min','TM3_max','TM3_min',
                    'TM4_max','TM4_min','TM5_max','TM5_min']
    temp['date'] = pd.to_datetime(temp['date'], errors='coerce')
    temp = temp.dropna(subset=['date']).set_index('date')
    return temp.apply(pd.to_numeric, errors='coerce')


def load_compensators(path):
    xl = pd.read_excel(path, sheet_name='Перечень сильфонных компенсатор', header=None)
    comps = xl.iloc[3:, [0, 2, 3, 4, 5, 7, 8, 9, 11, 12, 13]].copy()
    comps.columns = ['name','from_tk','to_tk','diameter','year',
                     'deltaL','L','T_mount','dT_100','dT_70','dT_20']
    comps = comps[comps['name'].notna() & comps['dT_100'].notna()].copy()
    comps['TM'] = comps['name'].apply(lambda x: TM_MAP.get(str(x)[:4], None))
    for col in ['year','dT_100','dT_70','dT_20','deltaL','L','T_mount']:
        comps[col] = pd.to_numeric(comps[col], errors='coerce')
    return comps.reset_index(drop=True)


def compute_results(comps, temp):
    results = []
    for _, row in comps.iterrows():
        tm = row['TM']
        year = int(row['year']) if pd.notna(row['year']) else None
        base = dict(name=row['name'], from_tk=row['from_tk'], to_tk=row['to_tk'],
                    diameter=row['diameter'], year=year, deltaL=row['deltaL'],
                    L=row['L'], T_mount=row['T_mount'], dT_100=row['dT_100'],
                    dT_70=row['dT_70'], dT_20=row['dT_20'], TM=tm)

        if tm not in TEMP_COLS or year is None:
            results.append({**base, 'cycles_100': None, 'cycles_70': None,
                            'cycles_20': None, 'days': 0, 'date_from': None,
                            'date_to': None, 'periods_100': [], 'periods_70': [],
                            'note': f'Нет данных температуры для {tm}'})
            continue

        max_col, min_col = TEMP_COLS[tm]
        series_max = temp[max_col][temp.index.year >= year]

        mask = temp[max_col].notna() & temp[min_col].notna()
        dT_daily = (temp[max_col] - temp[min_col]).where(mask).clip(lower=0)
        dT_daily = dT_daily[dT_daily.index.year >= year].dropna()

        if len(dT_daily) == 0:
            results.append({**base, 'cycles_100': 0, 'cycles_70': 0, 'cycles_20': 0,
                            'days': 0, 'date_from': None, 'date_to': None,
                            'periods_100': [], 'periods_70': [],
                            'note': f'Нет данных с {year} г.'})
            continue

        c100 = count_seasonal_cycles(series_max, row['T_mount'] + row['dT_100'])
        c70  = count_seasonal_cycles(series_max, row['T_mount'] + row['dT_70'])
        c20  = count_daily_cycles_carryover(dT_daily, row['dT_20'])

        periods_100 = find_seasonal_cycle_periods(series_max, row['T_mount'] + row['dT_100'])
        periods_70  = find_seasonal_cycle_periods(series_max, row['T_mount'] + row['dT_70'])

        results.append({
            **base, 'cycles_100': c100, 'cycles_70': c70, 'cycles_20': c20,
            'days': len(dT_daily),
            'date_from': dT_daily.index.min(), 'date_to': dT_daily.index.max(),
            'periods_100': periods_100, 'periods_70': periods_70,
            'note': ''
        })
    return results


def pct_color(val, limit):
    if val is None:
        return CLR_NODATA
    p = val / limit
    if p >= 1.0:
        return CLR_WARN
    if p >= 0.8:
        return CLR_CAUTION
    return CLR_OK


def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def write_cell(ws, row, col, value, bold=False, color=None, bg=None,
               align='left', fmt=None, border=True, size=9):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name='Arial', bold=bold,
                     color=('FF' + color) if color else 'FF000000', size=size)
    if bg:
        cell.fill = PatternFill('solid', start_color='FF' + bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = thin_border()


def build_xlsx(results, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Циклы СКУ'
    ws.freeze_panes = 'A4'

    ws.merge_cells('A1:T1')
    c = ws['A1']
    c.value = 'РАСЧЁТ ЦИКЛОВ СИЛЬФОННЫХ КОМПЕНСАТОРОВ'
    c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=12)
    c.fill = PatternFill('solid', start_color='FF' + CLR_HEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = [
        ('A','СКУ'), ('B','ТМ'), ('C','от ТК'), ('D','до ТК'), ('E','Ду, мм'),
        ('F','Год\nввода'), ('G','ΔL\nмм'), ('H','L тр.\nм'), ('I','Тм\n°C'),
        ('J','ΔТр 100%\n°C'), ('K','ΔТр 70%\n°C'), ('L','ΔТр 20%\n°C'),
        ('M','Данных\nдней'), ('N','Период\nот'), ('O','Период\nдо'),
        ('P', f'100%\n(лимит {LIMIT_100})'), ('Q', f'70%\n(лимит {LIMIT_70})'),
        ('R', f'20%\n(лимит {LIMIT_20})'), ('S','% от лимита\n100%'), ('T','Примечание'),
    ]
    for col_ltr, title in headers:
        ws.merge_cells(f'{col_ltr}2:{col_ltr}3')
        c = ws[f'{col_ltr}2']
        c.value = title
        c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
        c.fill = PatternFill('solid', start_color='FF' + CLR_SUBHDR)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 5

    col_w = [10,7,10,10,7,6,7,8,6,10,10,10,8,12,12,8,8,8,9,28]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, r in enumerate(results):
        row = idx + 4
        bg = CLR_ROW_A if idx % 2 == 0 else CLR_ROW_B
        def wc(col, val, **kw):
            write_cell(ws, row, col, val, bg=bg, **kw)

        wc(1, r['name'], bold=True); wc(2, r['TM']); wc(3, r['from_tk']); wc(4, r['to_tk'])
        wc(5, r['diameter'], align='center'); wc(6, r['year'], align='center')
        wc(7, r['deltaL'], align='right', fmt='0.0'); wc(8, r['L'], align='right', fmt='0.00')
        wc(9, r['T_mount'], align='right', fmt='0')
        wc(10, r['dT_100'], align='right', fmt='0.00'); wc(11, r['dT_70'], align='right', fmt='0.00')
        wc(12, r['dT_20'], align='right', fmt='0.00'); wc(13, r['days'], align='center')

        df_str = r['date_from'].strftime('%d.%m.%Y') if r['date_from'] else '—'
        dt_str = r['date_to'].strftime('%d.%m.%Y') if r['date_to'] else '—'
        wc(14, df_str, align='center'); wc(15, dt_str, align='center')

        for col_num, val, limit in [(16, r['cycles_100'], LIMIT_100),
                                     (17, r['cycles_70'], LIMIT_70),
                                     (18, r['cycles_20'], LIMIT_20)]:
            clr = pct_color(val, limit)
            disp = val if val is not None else '—'
            write_cell(ws, row, col_num, disp, bold=(val is not None and val >= limit),
                       color=clr, bg=bg, align='center')

        if r['cycles_100'] is not None:
            pct = round(r['cycles_100'] / LIMIT_100 * 100, 1)
            clr = pct_color(r['cycles_100'], LIMIT_100)
            write_cell(ws, row, 19, f"{pct}%", color=clr, bg=bg, align='center')
        else:
            wc(19, '—', align='center')

        wc(20, r['note'], color=CLR_NODATA if r['note'] else None)
        ws.row_dimensions[row].height = 16

    leg_row = len(results) + 5
    ws.cell(leg_row, 1).value = 'Легенда:'
    ws.cell(leg_row, 1).font = Font(name='Arial', bold=True, size=9)
    legend = [(CLR_WARN, 'Превышение лимита циклов'), (CLR_CAUTION, '>80% от лимита'),
              (CLR_OK, 'В пределах нормы'), (CLR_NODATA, 'Нет данных температуры')]
    for i, (clr, label) in enumerate(legend):
        c = ws.cell(leg_row + 1 + i, 1)
        c.fill = PatternFill('solid', start_color='FF' + clr)
        c.value = label
        c.font = Font(name='Arial', color='FFFFFFFF', size=9)

    # ── Лист 2: Методология ──────────────────────────────────────────────
    ws2 = wb.create_sheet('Методология')
    ws2.column_dimensions['A'].width = 95

    warn_row = 1
    ws2.merge_cells('A1:A1')
    c = ws2.cell(1, 1)
    c.value = ('⚠ я не уверен что это правильно ')
    c.font = Font(name='Arial', bold=True, color='FF000000', size=10)
    c.fill = PatternFill('solid', start_color='FF' + CLR_ASSUMPTION)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws2.row_dimensions[1].height = 110

    methodology = [
        ('', False, None, None),
        ('1. ИСХОДНЫЕ ДАННЫЕ', True, CLR_SUBHDR, 'FFFFFF'),
        ('  • Температурный файл: ежедневные Тмакс и Тмин подающего трубопровода для каждой ТМ', False, None, None),
        ('  • Файл компенсаторов: Тм (температура монтажа), ΔТр для 100%, 70%, 20% хода, год ввода', False, None, None),
        ('  • Используется только подающий трубопровод (не обратка)', False, None, None),
        ('', False, None, None),
        ('2. ЦИКЛЫ 100% И 70% (СЕЗОННЫЕ)', True, CLR_SUBHDR, 'FFFFFF'),
        ('  Гипотеза: сифон растягивается с прогревом сети (отопительный сезон) и', False, None, None),
        ('  возвращается в исходное положение летом — это редкое, протяжённое движение.', False, None, None),
        ('  Порог = Т_монтажа + ΔТр(100% или 70%)', False, None, None),
        ('  Цикл засчитывается когда суточный Tmax поднимается выше порога,', False, None, None),
        ('  а затем (через какое-то время) опускается обратно ниже порога — переход "туда-обратно".', False, None, None),
        ('', False, None, None),
        ('3. ЦИКЛЫ 20% (СУТОЧНЫЕ)', True, CLR_SUBHDR, 'FFFFFF'),
        ('  ΔT за сутки = Тмакс − Тмин (суточные быстрые колебания температуры трубы)', False, None, None),
        ('  Накопление с переносом остатка (carry-over), без привязки к положению сифона:', False, None, None),
        ('    available = остаток_вчера + ΔT_сегодня', False, None, None),
        ('    n_циклов = floor(available / ΔТр_20%)', False, None, None),
        ('    новый остаток = available − n × ΔТр_20%  (переносится на следующий день)', False, None, None),
        ('', False, None, None),
        ('4. ЛИМИТЫ (из паспортных данных компенсаторов, достоверны)', True, CLR_SUBHDR, 'FFFFFF'),
        (f'  100% полный ход: {LIMIT_100} циклов', False, None, None),
        (f'  70% хода: {LIMIT_70} циклов', False, None, None),
        (f'  20% хода: {LIMIT_20} циклов', False, None, None),
        ('', False, None, None),
        ('5. ОГРАНИЧЕНИЯ И ЗАМЕЧАНИЯ', True, CLR_SUBHDR, 'FFFFFF'),
        ('  • Три счётчика (100%, 70%, 20%) считаются НЕЗАВИСИМО, по разным принципам', False, None, None),
        ('  • Для ТМ-1 за 2007–2016 гг. есть только Тмакс (нет Тмин) → влияет на сезонный расчёт,', False, None, None),
        ('    но дни без Тмин исключены из суточного (20%) расчёта', False, None, None),
        ('  • СКУ6, СКУ7 — нет температурных данных в исходном файле, циклы не рассчитаны', False, None, None),
        ('  • Для СКУ1-23/24 порог 70% (ΔТр) выше порога 100% — особенность исходных паспортных', False, None, None),
        ('    данных (не ошибка расчёта); итоговые цифры по этим двум СКУ требуют отдельной проверки', False, None, None),
        ('', False, None, None),
    ]
    for i, (text, bold, bg, fg) in enumerate(methodology, 2):
        c = ws2.cell(i, 1, text)
        c.font = Font(name='Arial', bold=bold, color=('FF'+fg) if fg else 'FF000000', size=10)
        if bg:
            c.fill = PatternFill('solid', start_color='FF' + bg)
        ws2.row_dimensions[i].height = 16

    # ── Лист 3: Периоды циклов 100%/70% ──────────────────────────────────
    ws3 = wb.create_sheet('Периоды 100-70%')
    ws3.merge_cells('A1:E1')
    c = ws3['A1']
    c.value = 'ДАТЫ ПРЕВЫШЕНИЯ И ВОЗВРАТА ПОРОГОВ 100% И 70% (сезонные циклы)'
    c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=12)
    c.fill = PatternFill('solid', start_color='FF' + CLR_HEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 24

    headers3 = [('A','СКУ'), ('B','Уровень'), ('C','№ цикла'),
                ('D','Дата превышения порога\n(начало растяжения)'),
                ('E','Дата возврата ниже порога\n(конец цикла)')]
    for col_ltr, title in headers3:
        c = ws3[f'{col_ltr}2']
        c.value = title
        c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
        c.fill = PatternFill('solid', start_color='FF' + CLR_SUBHDR)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws3.row_dimensions[2].height = 28
    for col_ltr, w in zip('ABCDE', [12, 10, 9, 22, 22]):
        ws3.column_dimensions[col_ltr].width = w

    row3 = 3
    for r in results:
        any_periods = r.get('periods_100') or r.get('periods_70')
        if not any_periods:
            continue
        for level_name, periods, clr in [('100%', r.get('periods_100', []), CLR_WARN),
                                          ('70%',  r.get('periods_70', []),  CLR_CAUTION)]:
            for i, (d_up, d_down) in enumerate(periods, 1):
                write_cell(ws3, row3, 1, r['name'], bold=True)
                write_cell(ws3, row3, 2, level_name, align='center', color=clr, bold=True)
                write_cell(ws3, row3, 3, i, align='center')
                write_cell(ws3, row3, 4, d_up.strftime('%d.%m.%Y'), align='center')
                write_cell(ws3, row3, 5, d_down.strftime('%d.%m.%Y'), align='center')
                row3 += 1

    if row3 == 3:
        write_cell(ws3, 3, 1, 'Нет зафиксированных переходов порога 100%/70% ни у одного СКУ')

    wb.save(out_path)
    print(f'Saved: {out_path}')


def main(temp_path, comp_path, out_path):
    print('Загрузка температурных данных...')
    temp = load_temperatures(temp_path)
    print('Загрузка данных компенсаторов...')
    comps = load_compensators(comp_path)
    print(f'Найдено СКУ: {len(comps)}')
    results = compute_results(comps, temp)
    print('Формирование отчёта...')
    build_xlsx(results, out_path)

    print('\n═══ ИТОГ ═══')
    print(f"{'СКУ':<12}{'100%':>6}{'70%':>6}{'20%':>8}  Примечание")
    print('─'*60)
    for r in results:
        c100 = str(r['cycles_100']) if r['cycles_100'] is not None else '—'
        c70  = str(r['cycles_70'])  if r['cycles_70']  is not None else '—'
        c20  = str(r['cycles_20'])  if r['cycles_20']  is not None else '—'
        flag = ''
        if r['cycles_100'] is not None and r['cycles_100'] >= LIMIT_100:
            flag = '⚠ ЛИМИТ 100%!'
        elif r['cycles_70'] is not None and r['cycles_70'] >= LIMIT_70:
            flag = '⚠ ЛИМИТ 70%!'
        print(f"{r['name']:<12}{c100:>6}{c70:>6}{c20:>8}  {r['note'] or flag}")

        for level_name, periods in [('100%', r.get('periods_100', [])),
                                     ('70%',  r.get('periods_70', []))]:
            for i, (d_up, d_down) in enumerate(periods, 1):
                print(f"    └─ {level_name} цикл #{i}: {d_up.strftime('%d.%m.%Y')} → {d_down.strftime('%d.%m.%Y')}")


if __name__ == '__main__':
    if len(sys.argv) == 4:
        main(sys.argv[1], sys.argv[2], sys.argv[3])
    else:
        main('Данные_по_температурам_сетевой_воды_2__Автосохраненный_ (1).xlsx',
             'Приложение_к_протоколу.xlsx',
             'Результаты_циклов_СКУ.xlsx')